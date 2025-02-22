import sqlite3
import openpyxl
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s: %(message)s')

def dynamic_populate(file_path, db_path, table_name):
    """
    Importiert Daten aus einer Excel-Datei in die angegebene Tabelle,
    indem die Excel-Header mit den Datenbankspalten abgeglichen werden.
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        headers = [str(cell).strip().lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    except Exception as e:
        logging.error(f"Fehler beim Laden der Excel-Datei '{file_path}': {e}")
        return

    try:
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            cursor.execute(f"PRAGMA table_info({table_name})")
            db_columns = [row[1].strip().lower() for row in cursor.fetchall()]
            common = [col for col in db_columns if col in headers]
            if not common:
                logging.warning(f"Keine gemeinsamen Spalten in '{file_path}' und Tabelle '{table_name}'.")
                return
            header_index = {h: i for i, h in enumerate(headers)}
            placeholders = ", ".join(["?"] * len(common))
            sql = f"INSERT INTO {table_name} ({', '.join(common)}) VALUES ({placeholders})"
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data = [row[header_index[col]] if header_index.get(col) is not None and header_index[col] < len(row) else None for col in common]
                cursor.execute(sql, tuple(data))
            conn.commit()
            logging.info(f"Daten aus '{file_path}' in Tabelle '{table_name}' importiert.")
    except Exception as e:
        logging.error(f"Fehler beim Import in Tabelle '{table_name}': {e}")

if __name__ == "__main__":
    # Beispielaufruf:
    # dynamic_populate("deine_datei.xlsx", "instandhaltung.db", "deine_tabelle")
    pass

