import sqlite3
import openpyxl
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s: %(message)s')

def init_reference_data_from_excel(excel_file="Konfiguration.xlsx", db_path="instandhaltung.db"):
    """
    Importiert Referenzdaten aus der Excel-Konfiguration:
      - Jeder Blattname (außer Sonderblätter) wird als Abteilung genutzt.
      - Zeile 1: Anlagen; ab Zeile 2: Anlagenteile.
    """
    try:
        wb = openpyxl.load_workbook(excel_file)
        logging.info(f"Excel '{excel_file}' geladen.")
    except Exception as e:
        logging.error(f"Fehler beim Laden von '{excel_file}': {e}")
        return
    
    try:
        with sqlite3.connect(db_path) as conn:
            conn.execute("PRAGMA foreign_keys = ON;")
            cursor = conn.cursor()
            skip_sheets = {"General", "OptionalFields"}
            
            for sheet_name in wb.sheetnames:
                if sheet_name in skip_sheets:
                    continue
                sheet = wb[sheet_name]
                abt_name = sheet_name.strip()
                if not abt_name:
                    continue
                cursor.execute("INSERT OR IGNORE INTO abteilungen (name) VALUES (?)", (abt_name,))
                conn.commit()
                cursor.execute("SELECT abteilung_id FROM abteilungen WHERE name=?", (abt_name,))
                row = cursor.fetchone()
                if row is None:
                    continue
                abt_id = row[0]
                
                # Anlagen aus Zeile 1
                rows = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                if not rows:
                    continue
                anlagen = [str(cell).strip() for cell in rows[0] if cell]
                anlagen_ids = []
                for anlage in anlagen:
                    cursor.execute("INSERT OR IGNORE INTO anlagen (abteilung_id, name) VALUES (?, ?)", (abt_id, anlage))
                    conn.commit()
                    cursor.execute("SELECT anlage_id FROM anlagen WHERE abteilung_id=? AND name=?", (abt_id, anlage))
                    r = cursor.fetchone()
                    anlagen_ids.append(r[0] if r else None)
                
                # Anlagenteile ab Zeile 2
                for row_values in sheet.iter_rows(min_row=2, values_only=True):
                    for idx, cell in enumerate(row_values):
                        if cell and idx < len(anlagen_ids) and anlagen_ids[idx]:
                            anlagenteil = str(cell).strip()
                            if anlagenteil:
                                cursor.execute("INSERT OR IGNORE INTO anlagenteile (anlage_id, name) VALUES (?, ?)",
                                               (anlagen_ids[idx], anlagenteil))
                conn.commit()
                logging.info(f"Referenzdaten für Abteilung '{abt_name}' importiert.")
            logging.info("Alle Referenzdaten importiert.")
    except Exception as e:
        logging.error(f"Fehler beim Import der Referenzdaten: {e}")

def dynamic_populate(file_path, db_path, table_name):
    """
    Importiert Daten aus einer Excel-Datei in die angegebene Tabelle,
    indem die Excel-Header mit den Datenbankspalten abgeglichen werden.
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        headers = [str(cell).strip().lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    except Exception as e:
        logging.error(f"Fehler beim Laden der Excel-Datei '{file_path}': {e}")
        return

    try:
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            cursor.execute(f"PRAGMA table_info({table_name})")
            db_columns = [row[1].strip().lower() for row in cursor.fetchall()]
            common = [col for col in db_columns if col in headers]
            if not common:
                logging.warning(f"Keine gemeinsamen Spalten in '{file_path}' und Tabelle '{table_name}'.")
                return
            header_index = {h: i for i, h in enumerate(headers)}
            placeholders = ", ".join(["?"] * len(common))
            sql = f"INSERT INTO {table_name} ({', '.join(common)}) VALUES ({placeholders})"
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data = [row[header_index[col]] if header_index.get(col) is not None and header_index[col] < len(row) else None for col in common]
                cursor.execute(sql, tuple(data))
            conn.commit()
            logging.info(f"Daten aus '{file_path}' in Tabelle '{table_name}' importiert.")
    except Exception as e:
        logging.error(f"Fehler beim Import in Tabelle '{table_name}': {e}")

if __name__ == "__main__":
    # Beispielaufrufe:
    init_reference_data_from_excel()
    # dynamic_populate("Motorenlager_standardisiert.xlsx", "instandhaltung.db", "motoren")

