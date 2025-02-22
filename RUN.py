#!/usr/bin/env python
import sqlite3
import bcrypt
import openpyxl
import logging
import sys

# Logging-Konfiguration
logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s: %(message)s')

# Tabellen-Definitionen (wie in createtables.py)
TABLES = {
    "benutzer": """
        CREATE TABLE IF NOT EXISTS benutzer (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            password BLOB,
            role TEXT DEFAULT 'user'
        );
    """,
    "pannen": """
        CREATE TABLE IF NOT EXISTS pannen (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            abteilung TEXT,
            datum TEXT,
            schicht TEXT,
            name TEXT,
            anlage TEXT,
            anlagenteil TEXT,
            baugruppe TEXT,
            beschreibung TEXT,
            fehlerkategorie TEXT,
            fehlerursache TEXT,
            ausfallzeit TEXT,
            prioritaet TEXT,
            massnahme TEXT,
            melder TEXT
        );
    """,
    "motoren": """
        CREATE TABLE IF NOT EXISTS motoren (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            motornummer TEXT,
            im_sw TEXT,
            g TEXT,
            bs TEXT,
            firma TEXT,
            neu TEXT,
            typ TEXT,
            seriennummer TEXT,
            leistung TEXT,
            spannung TEXT,
            n1_min TEXT,
            n2_min TEXT,
            strom TEXT,
            cosinus_phi TEXT,
            lagerort TEXT,
            bemerkung TEXT
        );
    """,
    "ersatzteile": """
        CREATE TABLE IF NOT EXISTS ersatzteile (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            hersteller TEXT,
            typ TEXT,
            bauform TEXT,
            spannung TEXT,
            bestellnummer TEXT,
            lagerplatz TEXT,
            beschreibung TEXT,
            zusatz1 TEXT,
            zusatz2 TEXT,
            zusatz3 TEXT,
            bestand INTEGER DEFAULT 0,
            mindestbestand INTEGER DEFAULT 0,
            lieferant TEXT,
            einzelpreis REAL,
            waehrung TEXT
        );
    """,
    "wartungen": """
        CREATE TABLE IF NOT EXISTS wartungen (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            anlage TEXT,
            datum TEXT,
            pruefnotiz TEXT,
            wiederholung TEXT,
            erledigt INTEGER
        );
    """,
    "abteilungen": """
        CREATE TABLE IF NOT EXISTS abteilungen (
            abteilung_id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        );
    """,
    "anlagen": """
        CREATE TABLE IF NOT EXISTS anlagen (
            anlage_id INTEGER PRIMARY KEY AUTOINCREMENT,
            abteilung_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            FOREIGN KEY(abteilung_id) REFERENCES abteilungen(abteilung_id)
                ON DELETE CASCADE
                ON UPDATE CASCADE
        );
    """,
    "anlagenteile": """
        CREATE TABLE IF NOT EXISTS anlagenteile (
            anlagenteil_id INTEGER PRIMARY KEY AUTOINCREMENT,
            anlage_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            FOREIGN KEY(anlage_id) REFERENCES anlagen(anlage_id)
                ON DELETE CASCADE
                ON UPDATE CASCADE
        );
    """
}

def create_tables(db_path="instandhaltung.db"):
    """
    Erstellt bzw. aktualisiert alle notwendigen Tabellen in der SQLite-Datenbank
    und fügt Standardnutzer (admin und user) hinzu.
    """
    try:
        with sqlite3.connect(db_path) as conn:
            conn.execute("PRAGMA foreign_keys = ON;")
            cursor = conn.cursor()
            for table_name, query in TABLES.items():
                cursor.execute(query)
                logging.info(f"Tabelle '{table_name}' erstellt oder aktualisiert.")
            
            # Standardnutzer einfügen
            def insert_user(username, password, role):
                cursor.execute("SELECT COUNT(*) FROM benutzer WHERE username = ?", (username,))
                if cursor.fetchone()[0] == 0:
                    hashed = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt())
                    cursor.execute("INSERT INTO benutzer (username, password, role) VALUES (?, ?, ?)",
                                   (username, hashed, role))
                    logging.info(f"Benutzer '{username}' eingefügt.")
                else:
                    logging.info(f"Benutzer '{username}' existiert bereits.")
            
            insert_user("admin", "admin", "admin")
            insert_user("user", "user", "user")
            conn.commit()
            logging.info("Datenbanksetup abgeschlossen.")
    except Exception as e:
        logging.error(f"Fehler beim Erstellen der Tabellen: {e}")

def init_reference_data_from_excel(excel_file="Konfiguration.xlsx", db_path="instandhaltung.db"):
    """
    Importiert Abteilungen, Anlagen und Anlagenteile aus der Excel-Konfiguration.
      - Jeder Blattname (außer 'General' und 'OptionalFields') wird als Abteilung genutzt.
      - In Zeile 1 stehen die Anlagen, ab Zeile 2 die zugehörigen Anlagenteile.
    """
    try:
        wb = openpyxl.load_workbook(excel_file)
        logging.info(f"Excel-Datei '{excel_file}' erfolgreich geladen.")
    except Exception as e:
        logging.error(f"Fehler beim Öffnen der Excel-Datei '{excel_file}': {e}")
        return
    
    try:
        with sqlite3.connect(db_path) as conn:
            conn.execute("PRAGMA foreign_keys = ON;")
            cursor = conn.cursor()
            skip_sheets = {"General", "OptionalFields"}
            
            for sheet_name in wb.sheetnames:
                if sheet_name in skip_sheets:
                    logging.info(f"Blatt '{sheet_name}' wird übersprungen.")
                    continue
                sheet = wb[sheet_name]
                abteilungsname = sheet_name.strip()
                if not abteilungsname:
                    logging.warning("Leerer Blattname gefunden – wird übersprungen.")
                    continue
                cursor.execute("INSERT OR IGNORE INTO abteilungen (name) VALUES (?)", (abteilungsname,))
                conn.commit()
                cursor.execute("SELECT abteilung_id FROM abteilungen WHERE name=?", (abteilungsname,))
                row = cursor.fetchone()
                if row is None:
                    logging.error(f"Abteilung '{abteilungsname}' konnte nicht abgerufen werden.")
                    continue
                abteilung_id = row[0]
                
                # Anlagen aus Zeile 1 einlesen
                rows = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                if not rows:
                    logging.warning(f"Keine Anlagen im Blatt '{abteilungsname}' gefunden.")
                    continue
                anlagen_row = rows[0]
                anlagen_ids = []
                for cell in anlagen_row:
                    if cell is not None and str(cell).strip() != "":
                        anlage_name = str(cell).strip()
                        cursor.execute("INSERT OR IGNORE INTO anlagen (abteilung_id, name) VALUES (?, ?)",
                                       (abteilung_id, anlage_name))
                        conn.commit()
                        cursor.execute("SELECT anlage_id FROM anlagen WHERE abteilung_id=? AND name=?",
                                       (abteilung_id, anlage_name))
                        row_anlage = cursor.fetchone()
                        if row_anlage:
                            anlagen_ids.append(row_anlage[0])
                            logging.info(f"Anlage '{anlage_name}' in Abteilung '{abteilungsname}' hinzugefügt.")
                        else:
                            anlagen_ids.append(None)
                    else:
                        anlagen_ids.append(None)
                
                # Anlagenteile ab Zeile 2 einlesen
                for row_values in sheet.iter_rows(min_row=2, values_only=True):
                    for col_index, cell_value in enumerate(row_values):
                        if cell_value is not None and col_index < len(anlagen_ids) and anlagen_ids[col_index] is not None:
                            anlagenteil_name = str(cell_value).strip()
                            if anlagenteil_name:
                                cursor.execute("INSERT OR IGNORE INTO anlagenteile (anlage_id, name) VALUES (?, ?)",
                                               (anlagen_ids[col_index], anlagenteil_name))
                conn.commit()
                logging.info(f"Import für Abteilung '{abteilungsname}' abgeschlossen.")
            logging.info("Alle Referenzdaten wurden erfolgreich importiert.")
    except Exception as e:
        logging.error(f"Fehler beim Import der Referenzdaten: {e}")

def dynamic_populate(file_path, db_path, table_name):
    """
    Liest die Excel-Datei und überträgt alle Werte in die angegebene Tabelle.
    Es werden nur die Spalten befüllt, die sowohl in Excel als auch in der Datenbank vorhanden sind.
    Die erste Zeile der Excel-Datei enthält die Header.
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        excel_headers = [str(cell).strip().lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        logging.info(f"Excel-Header aus '{file_path}': {excel_headers}")
    except Exception as e:
        logging.error(f"Fehler beim Laden der Excel-Datei '{file_path}': {e}")
        return

    try:
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            cursor.execute(f"PRAGMA table_info({table_name})")
            db_columns = [row[1].strip().lower() for row in cursor.fetchall()]
            logging.info(f"Datenbankspalten in '{table_name}': {db_columns}")
            common_columns = [col for col in db_columns if col in excel_headers]
            if not common_columns:
                logging.warning(f"Keine gemeinsamen Spalten zwischen Excel und Tabelle '{table_name}' gefunden.")
                return
            header_index = {header: idx for idx, header in enumerate(excel_headers)}
            placeholders = ", ".join(["?"] * len(common_columns))
            sql = f"INSERT INTO {table_name} ({', '.join(common_columns)}) VALUES ({placeholders})"
            logging.info(f"SQL-Query: {sql}")
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data = [row[header_index[col]] if header_index.get(col) is not None and header_index[col] < len(row) else None for col in common_columns]
                try:
                    cursor.execute(sql, tuple(data))
                except Exception as e:
                    logging.error(f"Fehler beim Einfügen in Tabelle '{table_name}': {e}. Daten: {data}")
            conn.commit()
            logging.info(f"Daten aus '{file_path}' wurden erfolgreich in Tabelle '{table_name}' importiert.")
    except Exception as e:
        logging.error(f"Fehler beim Verbinden zur Datenbank '{db_path}': {e}")

def main():
    if len(sys.argv) < 2:
        print("Usage: python db_setup_import.py [create_tables | import_refdata | populate] [weitere Argumente]")
        sys.exit(1)
    
    action = sys.argv[1]
    db_path = "instandhaltung.db"  # Standard-Datenbankdatei

    if action == "create_tables":
        create_tables(db_path)
    elif action == "import_refdata":
        # Optional: Dateiname der Konfigurations-Excel als zweites Argument
        excel_file = sys.argv[2] if len(sys.argv) > 2 else "Konfiguration.xlsx"
        init_reference_data_from_excel(excel_file, db_path)
    elif action == "populate":
        # Für dynamic_populate: Erwarte Dateiname und Zieltabellenname als Argumente
        if len(sys.argv) < 4:
            print("Usage für populate: python db_setup_import.py populate <excel_file> <table_name>")
            sys.exit(1)
        file_path = sys.argv[2]
        table_name = sys.argv[3]
        dynamic_populate(file_path, db_path, table_name)
    else:
        print("Unbekannte Aktion. Bitte wähle: create_tables, import_refdata oder populate.")

if __name__ == "__main__":
    main()

